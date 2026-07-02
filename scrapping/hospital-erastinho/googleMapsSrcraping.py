import os
import time
from fpdf import FPDF
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook

# Função para garantir que o texto seja codificado corretamente
def safe_text(text):
    return text.encode('latin-1', 'replace').decode('latin-1')

# Capturar o tempo de início
start_time = time.time()

# Configurar o Selenium WebDriver
options = Options()
headless = False  # Altere para True para rodar sem interface gráfica
if headless:
    options.add_argument('--headless')

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# URL do site com as avaliações
url = 'https://www.google.com/maps/place/Hospital+Erastinho/@-25.452814,-49.2411939,1074m/data=!3m1!1e3!4m8!3m7!1s0x94dce5124cb3e99b:0xdb43625e49c502d9!8m2!3d-25.4528189!4d-49.238619!9m1!1b1!16s%2Fg%2F11h5mp07v2?entry=ttu&g_ep=EgoyMDI2MDYyOC4wIKXMDSoASAFQAw%3D%3D'
driver.get(url)

# Aguardar carregamento da página
WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.CLASS_NAME, 'jftiEf'))
)

limite_avaliacoes = 360  # Defina o limite de avaliações desejado
avaliacoes = []

# Função para carregar avaliações
def carregar_avaliacoes():
    try:
        scrollable_div = driver.find_element(By.CSS_SELECTOR, 'div.m6QErb.DxyBCb.kA9KIf.dS8AEf')
        last_height = driver.execute_script("return arguments[0].scrollHeight;", scrollable_div)

        while len(avaliacoes) < limite_avaliacoes:
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", scrollable_div)
            time.sleep(2)
            new_height = driver.execute_script("return arguments[0].scrollHeight;", scrollable_div)
            if new_height == last_height:
                break
            last_height = new_height

            # Capturar novas avaliações
            novas_avaliacoes = driver.find_elements(By.CLASS_NAME, 'jftiEf')
            for avaliacao in novas_avaliacoes:
                comentario_texto = avaliacao.text.strip()
                if comentario_texto and comentario_texto not in avaliacoes_textos:
                    avaliacoes.append(avaliacao)
                    avaliacoes_textos.add(comentario_texto)
                    if len(avaliacoes) >= limite_avaliacoes:
                        break
            if len(avaliacoes) >= limite_avaliacoes:
                break
    except Exception as e:
        print(f"Erro ao carregar avaliações: {e}")

# Armazenar apenas textos únicos das avaliações
avaliacoes_textos = set()

# Capturar avaliações
carregar_avaliacoes()

# Criar PDF e Excel
pdf = FPDF()
pdf.set_auto_page_break(auto=True, margin=15)
pdf.add_page()
pdf.set_font("Arial", size=12)

wb = Workbook()
ws = wb.active
ws.title = "Avaliações"
ws.append(["Nome", "Nota", "Data", "Comentário", "Resposta da Empresa"])

for avaliacao in avaliacoes:
    try:
        try:
            mais_button = avaliacao.find_element(By.CLASS_NAME, 'w8nwRe.kyuRq')
            mais_button.click()
            time.sleep(1)
        except:
            pass

        nome = avaliacao.find_element(By.CLASS_NAME, 'd4r55').text.strip()
    except:
        nome = 'Nome não encontrado'

    try:
        nota = avaliacao.find_element(By.CLASS_NAME, 'kvMYJc').get_attribute('aria-label')
    except:
        nota = 'Nota não encontrada'

    try:
        comentario = avaliacao.find_element(By.CLASS_NAME, 'MyEned').text.strip()
    except:
        comentario = 'Comentário não encontrado'

    try:
        data_comentario = avaliacao.find_element(By.CLASS_NAME, 'rsqaWe').text.strip()
    except:
        data_comentario = 'Data não encontrada'

    try:
        resposta_empresa = avaliacao.find_element(By.XPATH, './/div[contains(@class, "wiI7pd")]').text.strip()
    except:
        resposta_empresa = 'Resposta não encontrada'

    pdf.cell(200, 10, txt=safe_text(f"Nome: {nome}"), ln=True, align='L')
    pdf.cell(200, 10, txt=safe_text(f"Nota: {nota}"), ln=True, align='L')
    pdf.cell(200, 10, txt=safe_text(f"Data: {data_comentario}"), ln=True, align='L')
    pdf.multi_cell(200, 10, txt=safe_text(f"Comentário: {comentario}"), align='L')
    pdf.multi_cell(200, 10, txt=safe_text(f"Resposta da empresa: {resposta_empresa}"), align='L')
    pdf.cell(200, 10, txt=safe_text('-' * 40), ln=True, align='L')

    ws.append([nome, nota, data_comentario, comentario, resposta_empresa])

# Salvar arquivos

#pequena alteração para salvar os arquivos em uma pasta específica dentro do próprio programa 
#posteriormente pretendo criar uma inteligencia artificial que analise os dados e gere gráficos e insights a partir das avaliações coletadas
base_path = os.path.dirname(os.path.abspath(__file__))
documents_path = os.path.join(base_path, "avaliacoes")
#eh o mesmo esquema so que agora ele cria a pasta "avaliacoes" dentro do diretório do proprio programa caso ela não exista

os.makedirs(documents_path, exist_ok=True)

pdf_path = os.path.join(documents_path, "avaliacoes.pdf")
pdf.output(pdf_path)
print(f"PDF salvo com sucesso em: {pdf_path}")

excel_path = os.path.join(documents_path, "avaliacoes.xlsx")
wb.save(excel_path)
print(f"Excel salvo com sucesso em: {excel_path}")

# Fechar o navegador
driver.quit()

# Capturar o tempo de término
end_time = time.time()

# Calcular o tempo de execução
execution_time = end_time - start_time

# Converter o tempo de execução para minutos e segundos
minutes, seconds = divmod(execution_time, 60)
print(f"Tempo de execução do script: {int(minutes)} minutos e {seconds:.2f} segundos")